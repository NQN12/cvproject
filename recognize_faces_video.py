import argparse
import cv2
import os
import time
import datetime
import face_recognition
from openpyxl import Workbook, load_workbook

# Parse command-line arguments
ap = argparse.ArgumentParser()
ap.add_argument("-o", "--output", required=True, help="path to output directory")
ap.add_argument("--excel", type=str, default="faces.xlsx", help="path to the Excel file")
ap.add_argument("-d", "--detection_method", type=str, default="hog", help="face detection model to use: cnn or hog")
args = vars(ap.parse_args())

# Create output directory if it doesn't exist
if not os.path.exists(args["output"]):
    os.makedirs(args["output"])

# Set up Excel file
if not os.path.exists(args["excel"]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Timestamp", "Name", "Image Path"])
    workbook.save(args["excel"])
else:
    workbook = load_workbook(args["excel"])
    sheet = workbook.active

# Load the known faces and encodings
print("[INFO] loading encodings...")
with open('encodings.pickle', "rb") as file:
    data = pickle.load(file)

# Initialize webcam
video = cv2.VideoCapture(0)
writer = None
time.sleep(2.0)

# Timer and counter setup
detection_intervals = {}
CONSECUTIVE_DETECTION_THRESHOLD = 5
DETECTION_INTERVAL = 1  # seconds

# Main loop
while True:
    ret, frame = video.read()
    if not ret:
        break

    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    rgb = cv2.resize(rgb, (750, int(frame.shape[0] * 750 / frame.shape[1])))
    scale = frame.shape[1] / float(rgb.shape[1])

    boxes = face_recognition.face_locations(rgb, model=args["detection_method"])
    encodings = face_recognition.face_encodings(rgb, boxes)
    
    current_detections = []

    for encoding in encodings:
        matches = face_recognition.compare_faces(data["encodings"], encoding)
        name = "Unknown"
        if True in matches:
            matchedIdxs = [i for (i, b) in enumerate(matches) if b]
            counts = {}
            for i in matchedIdxs:
                name = data["names"][i]
                counts[name] = counts.get(name, 0) + 1
            name = max(counts, key=counts.get)
        
        current_detections.append(name)

    for name in current_detections:
        if name not in detection_intervals:
            detection_intervals[name] = 0
        detection_intervals[name] += 1
    
    # Check for people who have been detected for the required amount of time
    for name, count in list(detection_intervals.items()):
        if count >= CONSECUTIVE_DETECTION_THRESHOLD:
            timestamp = datetime.datetime.now()
            filename = f"{timestamp.strftime('%Y%m%d_%H%M%S%f')}.png"
            file_path = os.path.join(args["output"], filename)
            cv2.imwrite(file_path, frame)
            sheet.append([timestamp.strftime('%Y-%m-%d %H:%M:%S'), name, file_path])
            workbook.save(args["excel"])
            del detection_intervals[name]  # Reset the counter after writing to the Excel file

    # Draw bounding boxes and names on the frame
    for ((top, right, bottom, left), name) in zip(boxes, current_detections):
        top = int(top * scale)
        right = int(right * scale)
        bottom = int(bottom * scale)
        left = int(left * scale)

        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        y = top - 15 if top - 15 > 15 else top + 15

        cv2.putText(frame, name, (left, y), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 255, 0), 2)

    cv2.imshow("video", frame)
    key = cv2.waitKey(1) & 0xFF
    if key == ord("q"):
        break

    time.sleep(DETECTION_INTERVAL)

print("[INFO] Exiting and cleaning up.")
video.release()
cv2.destroyAllWindows()
workbook.save(args["excel"])
workbook.close()
