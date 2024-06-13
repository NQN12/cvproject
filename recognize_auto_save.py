import argparse
import cv2
import os
import time
import datetime
import face_recognition
from openpyxl import Workbook, load_workbook
import pickle
import numpy as np

# Parse command-line arguments
ap = argparse.ArgumentParser()
ap.add_argument("-o", "--output", required=True, help="path to output directory")
ap.add_argument("--excel", type=str, default="faces.xlsx", help="path to the Excel file")
ap.add_argument("--encodings", required=True, help="path to the serialized db of facial encodings")
ap.add_argument("-d", "--detection_method", type=str, default="hog", help="face detection model to use: cnn or hog")
args = vars(ap.parse_args())

# Create output directory if it doesn't exist
if not os.path.exists(args["output"]):
    os.makedirs(args["output"])

# Set up Excel file
if not os.path.exists(args["excel"]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Timestamp", "Name", "Image Path", "Confidence"])
    workbook.save(args["excel"])
else:
    workbook = load_workbook(args["excel"])
    sheet = workbook.active

# Load the known faces and encodings
print("[INFO] loading encodings...")
with open(args["encodings"], "rb") as file:
    data = pickle.load(file)

# Initialize webcam
video = cv2.VideoCapture(0)
writer = None
time.sleep(2.0)

# Timer and counter setup
detection_intervals = {}
consecutive_detections = {}
last_captured_times = {}
CONSECUTIVE_DETECTION_THRESHOLD = 5
DETECTION_INTERVAL = 1  # seconds
MIN_TIME_BETWEEN_CAPTURES = 120  # seconds

# Main loop
while True:
    ret, frame = video.read()
    if not ret:
        break

    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    rgb = cv2.resize(rgb, (750, int(frame.shape[0] * 750 / frame.shape[1])))
    scale = frame.shape[1] / float(rgb.shape[1])

    boxes = face_recognition.face_locations(rgb, model=args["detection_method"])
    encodings = face_recognition.face_encodings(rgb, boxes)
    
    current_detections = []

    for encoding in encodings:
        matches = face_recognition.compare_faces(data["encodings"], encoding)
        face_distances = face_recognition.face_distance(data["encodings"], encoding)
        best_match_index = np.argmin(face_distances)
        name = "Unknown"
        confidence = 0

        if matches[best_match_index]:
            name = data["names"][best_match_index]
            confidence = 1 - face_distances[best_match_index]  # Confidence score

        current_detections.append((name, confidence))

    # Update detection intervals and consecutive detections
    for (name, confidence) in current_detections:
        if name not in detection_intervals:
            detection_intervals[name] = 0
        detection_intervals[name] += 1

        if name not in consecutive_detections:
            consecutive_detections[name] = 0
        consecutive_detections[name] += 1

        current_time = datetime.datetime.now()

        if name not in last_captured_times or (current_time - last_captured_times[name]).total_seconds() > MIN_TIME_BETWEEN_CAPTURES:
            if consecutive_detections[name] >= CONSECUTIVE_DETECTION_THRESHOLD:
                filename = f"{current_time.strftime('%Y%m%d_%H%M%S%f')}.png"
                file_path = os.path.join(args["output"], filename)
                cv2.imwrite(file_path, frame)
                sheet.append([current_time.strftime('%Y-%m-%d %H:%M:%S'), name, file_path, confidence])
                workbook.save(args["excel"])
                consecutive_detections[name] = 0  # Reset the counter after writing to the Excel file
                last_captured_times[name] = current_time  # Update the last captured time

    # Draw bounding boxes, names, and confidence scores on the frame
    for ((top, right, bottom, left), (name, confidence)) in zip(boxes, current_detections):
        top = int(top * scale)
        right = int(right * scale)
        bottom = int(bottom * scale)
        left = int(left * scale)

        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        y = top - 15 if top - 15 > 15 else top + 15
        text = f"{name}: {confidence+0.1:.2f}"
        cv2.putText(frame, text, (left, y), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 255, 0), 2)

    cv2.imshow("video", frame)
    key = cv2.waitKey(1) & 0xFF
    if key == ord("q"):
        break

    time.sleep(DETECTION_INTERVAL)

print("[INFO] Exiting and cleaning up.")
video.release()
cv2.destroyAllWindows()
workbook.save(args["excel"])
workbook.close()
